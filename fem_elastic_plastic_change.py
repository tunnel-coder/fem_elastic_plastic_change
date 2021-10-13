# FEM
import numpy as np

# создание окна
from tkinter import *
import tkinter.ttk

import matplotlib.ticker as ticker
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment


def geometry_matrix_element(B, xi, xj, xm, zi, zj, zm):
    dlt = xi * (zj - zm) + xj * (zm - zi) + xm * (zi - zj)
    B[0, 0] = (zm - zj) / dlt
    B[0, 1] = 0
    B[0, 2] = (zi - zm) / dlt
    B[0, 3] = 0
    B[0, 4] = (zj - zi) / dlt
    B[0, 5] = 0
    B[1, 0] = 0
    B[1, 1] = (xj - xm) / dlt
    B[1, 2] = 0
    B[1, 3] = (xm - xi) / dlt
    B[1, 4] = 0
    B[1, 5] = (xi - xj) / dlt
    B[2, 0] = B[1, 1]
    B[2, 1] = B[0, 0]
    B[2, 2] = B[1, 3]
    B[2, 3] = B[0, 2]
    B[2, 4] = B[1, 5]
    B[2, 5] = B[0, 4]
    BT = np.transpose(B)
    return BT


def elastic_matrix_element(D, E, nu):
    Enu = E / (1 + nu)
    D[0, 0] = (1 - nu) / (1 - 2 * nu) * Enu
    D[0, 1] = nu / (1 - 2 * nu) * Enu
    D[0, 2] = 0
    D[1, 0] = D[0, 1]
    D[1, 1] = D[0, 0]
    D[1, 2] = 0
    D[2, 0] = 0
    D[2, 1] = 0
    D[2, 2] = 0.5 * Enu
    return D


def plastic_matrix_element(k, it, sx, sz, txz, kst, nu, fi, D, pl, E):
    if pl[k - 1] == 0:
        return D
    if it == 1:
        sxp = sx[kst - 2, k - 1]
        szp = sz[kst - 2, k - 1]
        txzp = txz[kst - 2, k - 1]
    else:
        sxp = sx[kst - 1, k - 1]
        szp = sz[kst - 1, k - 1]
        txzp = txz[kst - 1, k - 1]

    a0 = szp - sxp
    a1 = np.sqrt((szp - sxp) ** 2 + 4 * txzp ** 2)
    p1 = -a0 / a1 - np.sin(fi) / (1 - 2 * nu)
    p2 = a0 / a1 - np.sin(fi) / (1 - 2 * nu)
    p3 = 2 * txzp / a1
    q0 = 2 * (1 + np.sin(fi) ** 2 / (1 - 2 * nu))
    Enu = E / (1 + nu)
    D[0, 0] = D[0, 0] - p1 * p1 / q0 * Enu
    D[0, 1] = D[0, 1] - p1 * p2 / q0 * Enu
    D[0, 2] = D[0, 2] - p1 * p3 / q0 * Enu
    D[1, 0] = D[0, 1]
    D[1, 1] = D[1, 1] - p2 * p2 / q0 * Enu
    D[1, 2] = D[1, 2] - p2 * p3 / q0 * Enu
    D[2, 0] = D[0, 2]
    D[2, 1] = D[1, 2]
    D[2, 2] = D[2, 2] - p3 * p3 / q0 * Enu
    return D



def stiffness_matrix_local(BT, D, B, xi, xj, xm, zi, zj, zm):
    dlt = xi * (zj - zm) + xj * (zm - zi) + xm * (zi - zj)
    dlt = np.abs(dlt) / 2
    BTD = np.dot(BT, D)
    BTDB = np.dot(BTD, B)
    kloc = BTDB * dlt
    return kloc


def stiffness_matrix_total(kglob, ki, kj, kloc, km):
    nk1 = 0
    nk2 = 0
    for i1 in range(0, 6):
        if 0 <= i1 <= 1:
            nk1 = ki * 2 + i1 - 2
        elif 2 <= i1 <= 3:
            nk1 = kj * 2 + i1 - 4
        elif 4 <= i1 <= 5:
            nk1 = km * 2 + i1 - 6
        for j1 in range(0, 6):
            if 0 <= j1 <= 1:
                nk2 = ki * 2 + j1 - 2
            elif 2 <= j1 <= 3:
                nk2 = kj * 2 + j1 - 4
            elif 4 <= j1 <= 5:
                nk2 = km * 2 + j1 - 6
            kglob[nk1, nk2] = kglob[nk1, nk2] + kloc[i1, j1]
    return kglob


def strains(du, B, ki, kj, km, k, dex, dez, dexz, kst, ex, ez, exz):
    dex[k - 1] = (du[2 * ki - 2] * B[0, 0] + du[2 * kj - 2] * B[0, 2] + du[2 * km - 2] * B[0, 4])
    dez[k - 1] = (du[2 * ki - 1] * B[1, 1] + du[2 * kj - 1] * B[1, 3] + du[2 * km - 1] * B[1, 5])
    dexz[k - 1] = (du[2 * ki - 2] * B[2, 0] + du[2 * kj - 2] * B[2, 2] + du[2 * km - 2] * B[2, 4]
                   + du[2 * ki - 1] * B[2, 1] + du[2 * kj - 1] * B[2, 3] + du[2 * km - 1] * B[2, 5])
    if kst == 1:
        ex[kst - 1, k - 1] = dex[k - 1]
        ez[kst - 1, k - 1] = dez[k - 1]
        exz[kst - 1, k - 1] = dexz[k - 1]
    else:
        ex[kst - 1, k - 1] = ex[kst - 2, k - 1] + dex[k - 1]
        ez[kst - 1, k - 1] = ez[kst - 2, k - 1] + dez[k - 1]
        exz[kst - 1, k - 1] = exz[kst - 2, k - 1] + dexz[k - 1]
    return dex, dez, dexz, ex, ez, exz


def stresses(D, dex, dez, dexz, k, dsx, dsz, dtxz, s1, s3, kst, sx, sz, txz):
    dsx[k - 1] = D[0, 0] * dex[k - 1] + D[0, 1] * dez[k - 1] + D[0, 2] * dexz[k - 1]
    dsz[k - 1] = D[1, 0] * dex[k - 1] + D[1, 1] * dez[k - 1] + D[1, 2] * dexz[k - 1]
    dtxz[k - 1] = D[2, 0] * dex[k - 1] + D[2, 1] * dez[k - 1] + D[2, 2] * dexz[k - 1]
    if kst == 1:
        sx[kst - 1, k - 1] = dsx[k - 1]
        sz[kst - 1, k - 1] = dsz[k - 1]
        txz[kst - 1, k - 1] = dtxz[k - 1]
    else:
        sx[kst - 1, k - 1] = sx[kst - 2, k - 1] + dsx[k - 1]
        sz[kst - 1, k - 1] = sz[kst - 2, k - 1] + dsz[k - 1]
        txz[kst - 1, k - 1] = txz[kst - 2, k - 1] + dtxz[k - 1]

    s1[kst - 1, k - 1] = (sx[kst - 1, k - 1] + sz[kst - 1, k - 1]) / 2 + np.sqrt((sx[kst - 1, k - 1]
                                                                                  - sz[kst - 1, k - 1]) ** 2 / 4
                                                                                 + txz[kst - 1, k - 1] ** 2)
    s3[kst - 1, k - 1] = (sx[kst - 1, k - 1] + sz[kst - 1, k - 1]) / 2 - np.sqrt((sx[kst - 1, k - 1]
                                                                                  - sz[kst - 1, k - 1]) ** 2 / 4
                                                                                 + txz[kst - 1, k - 1] ** 2)
    return dex, dez, dexz, sx, sz, txz, s1, s3


def stress_average(k, dex, dez, dexz, D, dsx, dsz, dtxz, kst, sx, sz, txz, s1, s3):
    dex[k - 1] = (dex[k - 1] + dex[k - 2]) / 2
    dez[k - 1] = (dez[k - 1] + dez[k - 2]) / 2
    dexz[k - 1] = (dexz[k - 1] + dexz[k - 2]) / 2

    dsx[k - 1] = D[0, 0] * dex[k - 1] + D[0, 1] * dez[k - 1] + D[0, 2] * dexz[k - 1]
    dsz[k - 1] = D[1, 0] * dex[k - 1] + D[1, 1] * dez[k - 1] + D[1, 2] * dexz[k - 1]
    dtxz[k - 1] = D[2, 0] * dex[k - 1] + D[2, 1] * dez[k - 1] + D[2, 2] * dexz[k - 1]

    if kst == 1:
        sx[kst - 1, k - 1] = dsx[k - 1]
        sx[kst - 1, k - 2] = dsx[k - 1]
        sz[kst - 1, k - 1] = dsz[k - 1]
        sz[kst - 1, k - 2] = dsz[k - 1]
        txz[kst - 1, k - 1] = dtxz[k - 1]
        txz[kst - 1, k - 2] = dtxz[k - 1]
    else:
        sx[kst - 1, k - 1] = sx[kst - 2, k - 1] + dsx[k - 1]
        sx[kst - 1, k - 2] = sx[kst - 1, k - 1]
        sz[kst - 1, k - 1] = sz[kst - 2, k - 1] + dsz[k - 1]
        sz[kst - 1, k - 2] = sz[kst - 1, k - 1]
        txz[kst - 1, k - 1] = txz[kst - 2, k - 1] + dtxz[k - 1]
        txz[kst - 1, k - 2] = txz[kst - 1, k - 1]

    s1[kst - 1, k - 1] = (sx[kst - 1, k - 1] + sz[kst - 1, k - 1]) / 2 \
                         + np.sqrt((sx[kst - 1, k - 1] - sz[kst - 1, k - 1]) ** 2 / 4 + txz[kst - 1, k - 1] ** 2)
    s3[kst - 1, k - 1] = (sx[kst - 1, k - 1] + sz[kst - 1, k - 1]) / 2 \
                         - np.sqrt((sx[kst - 1, k - 1] - sz[kst - 1, k - 1]) ** 2 / 4 + txz[kst - 1, k - 1] ** 2)
    s1[kst - 1, k - 2] = s1[kst - 1, k - 1]
    s3[kst - 1, k - 2] = s3[kst - 1, k - 1]
    return dex, dez, dexz, sx, sz, txz, s1, s3


def funcplast(sxFp, szFp, txzFp, fi, c):
    return np.sqrt((szFp - sxFp) ** 2 + 4 * txzFp ** 2) - ((szFp + sxFp) * np.sin(fi) + 2 * c * np.cos(fi))


def function_plastic(sx, sz, txz, kst, k, fi, c, it, Fp, FF, pl, npl, nu, xi, xj, xm, zi, zj, zm, ki, kj, km, B, dR,
                     dex, dez, dexz, ex, ez, exz, E, plastic, average):
    if plastic == 0:
        return sx, sz, txz, ex, ez, exz, dR, npl, Fp, FF, dex, dez, dexz
    Fpst = funcplast(sx[kst - 1, k - 1], sz[kst - 1, k - 1], txz[kst - 1, k - 1], fi, c)
    Fp[kst - 1, k - 1] = Fpst
    FF[kst - 1, k - 1, it - 1] = Fpst
    if average == 1:
        Fp[kst - 1, k - 2] = Fp[kst - 1, k - 1]
        FF[kst - 1, k - 2, it - 1] = Fpst
    if Fpst < 0:
        return sx, sz, txz, ex, ez, exz, dR, npl, Fp, FF, dex, dez, dexz
    if pl[k - 1] == 1:
        return sx, sz, txz, ex, ez, exz, dR, npl, Fp, FF, dex, dez, dexz
    pl[k - 1] = 1
    npl[kst - 1, it - 1] = npl[kst - 1, it - 1] + 1
    if average == 1:
        pl[k - 2] = 8
        npl[kst - 1, it - 1] = npl[kst - 1, it - 1] + 1
    sxF = 0
    szF = 0
    txzF = 0
    for kFst in np.arange(0, 1.02, 0.02):
        sxF = sx[kst - 2, k - 1] + kFst * (sx[kst - 1, k - 1] - sx[kst - 2, k - 1])
        szF = sz[kst - 2, k - 1] + kFst * (sz[kst - 1, k - 1] - sz[kst - 2, k - 1])
        txzF = txz[kst - 2, k - 1] + kFst * (txz[kst - 1, k - 1] - txz[kst - 2, k - 1])
        Fpst = funcplast(sxF, szF, txzF, fi, c)
        if Fpst > 0:
            sx[kst - 1, k - 1] = sxF
            sz[kst - 1, k - 1] = szF
            txz[kst - 1, k - 1] = txzF
            Fp[kst - 1, k - 1] = Fpst
            FF[kst - 1, k - 1, it - 1] = Fpst
            if average == 1:
                Fp[kst - 1, k - 2] = Fp[kst - 1, k - 1]
                FF[kst - 1, k - 2, it - 1] = Fpst
                sx[kst - 1, k - 2] = sxF
                sz[kst - 1, k - 2] = szF
                txz[kst - 1, k - 2] = txzF
            break
    a0 = szF - sxF
    a1 = np.sqrt((szF - sxF) ** 2 + 4 * txzF ** 2)
    p1 = -a0 / a1 - np.sin(fi) / (1 - 2 * nu)
    p2 = a0 / a1 - np.sin(fi) / (1 - 2 * nu)
    p3 = 2 * txzF / a1
    f1 = -a0 / a1 - np.sin(fi)
    f2 = a0 / a1 - np.sin(fi)
    f3 = 4 * txzF / a1
    q0 = 2 * (1 + np.sin(fi) ** 2 / (1 - 2 * nu))
    a6 = f1 * (sxF - sx[kst - 2, k - 1]) + f2 * (szF - sz[kst - 2, k - 1]) + f3 * (txzF - txz[kst - 2, k - 1])
    a1 = p1 * a6 / q0
    a2 = p2 * a6 / q0
    a3 = p3 * a6 / q0

    dlt = np.abs(xi * (zj - zm) + xj * (zm - zi) + xm * (zi - zj)) / 2

    dR[2 * ki - 2] = dR[2 * ki - 2] - dlt * (a1 * B[0, 0] + a3 * B[2, 0])
    dR[2 * ki - 1] = dR[2 * ki - 1] - dlt * (a2 * B[1, 1] + a3 * B[2, 1])
    dR[2 * kj - 2] = dR[2 * kj - 2] - dlt * (a1 * B[0, 2] + a3 * B[2, 2])
    dR[2 * kj - 1] = dR[2 * kj - 1] - dlt * (a2 * B[1, 3] + a3 * B[2, 3])
    dR[2 * km - 2] = dR[2 * km - 2] - dlt * (a1 * B[0, 4] + a3 * B[2, 4])
    dR[2 * km - 1] = dR[2 * km - 1] - dlt * (a2 * B[1, 5] + a3 * B[2, 5])

    sx[kst - 1, k - 1] = sx[kst - 2, k - 1] + a1
    sz[kst - 1, k - 1] = sz[kst - 2, k - 1] + a2
    txz[kst - 1, k - 1] = txz[kst - 2, k - 1] + a3

    Fpst = funcplast(sx[kst - 1, k - 1], sz[kst - 1, k - 1], txz[kst - 1, k - 1], fi, c)
    Fp[kst - 1, k - 1] = Fpst
    FF[kst - 1, k - 1, it - 1] = Fpst

    Enu = E / (1 + nu)
    dex[k - 1] = (a1 * (1 - nu) - a2 * nu) / Enu
    dez[k - 1] = (a2 * (1 - nu) - a1 * nu) / Enu
    dexz[k - 1] = 2 * a3 / Enu
    ex[kst - 1, k - 1] = ex[kst - 2, k - 1] + dex[k - 1]
    ez[kst - 1, k - 1] = ez[kst - 2, k - 1] + dez[k - 1]
    exz[kst - 1, k - 1] = exz[kst - 2, k - 1] + dexz[k - 1]

    if average == 1:
        sx[kst - 1, k - 2] = sx[kst - 1, k - 1]
        sz[kst - 1, k - 2] = sz[kst - 1, k - 1]
        txz[kst - 1, k - 2] = txz[kst - 1, k - 1]
        Fp[kst - 1, k - 2] = Fp[kst - 1, k - 1]
        FF[kst - 1, k - 2, it - 1] = Fpst
        dex[k - 1] = dex[k - 2]
        dez[k - 1] = dez[k - 2]
        dexz[k - 1] = dexz[k - 2]
        ex[kst - 1, k - 1] = ex[kst - 1, k - 2]
        ez[kst - 1, k - 1] = ez[kst - 1, k - 2]
        exz[kst - 1, k - 1] = exz[kst - 1, k - 2]
    return sx, sz, txz, ex, ez, exz, dR, npl, Fp, FF, dex, dez, dexz


class Application(Frame):
    def __init__(self, master):
        super(Application, self).__init__(master)
        self.var = IntVar()
        self.radioValue = IntVar()
        self.grid()
        self.create_widgets()

    def create_widgets(self):
        # Elastic_modulus
        Label(self, text="Elastic modulus(E)", font="Helvetica 9 italic").grid(row=0, column=0, sticky='w')
        self.E_ent = Entry(self, width=5, justify=CENTER)
        self.E_ent.grid(row=0, column=1)
        self.E_ent.insert(0, "10000")
        Label(self, text="kPa", font="Helvetica 9 italic").grid(row=0, column=2, sticky='w')

        # Poisson's_ratio
        Label(self, text="Poisson's ratio(ν)", font="Helvetica 9 italic").grid(row=1, column=0, sticky='w')
        self.nu_ent = Entry(self, width=5, justify=CENTER)
        self.nu_ent.grid(row=1, column=1)
        self.nu_ent.insert(0, "0.3")

        # Cohesion
        Label(self, text="Cohesion(c)", font="Helvetica 9 italic").grid(row=2, column=0, sticky='w')
        self.c_ent = Entry(self, width=5, justify=CENTER)
        self.c_ent.grid(row=2, column=1)
        self.c_ent.insert(0, "10")
        Label(self, text="kPa", font="Helvetica 9 italic").grid(row=2, column=2, sticky='w')

        # Frictional_angle
        Label(self, text="Frictional angle(φ)", font="Helvetica 9 italic").grid(row=3, column=0, sticky='w')
        self.fi_ent = Entry(self, width=5, justify=CENTER)
        self.fi_ent.grid(row=3, column=1)
        self.fi_ent.insert(0, "30")
        Label(self, text="°").grid(row=3, column=2, sticky='w')

        tkinter.ttk.Separator(self, orient=HORIZONTAL).grid(column=0, row=4, columnspan=3, sticky='we')

        # Pressure
        Label(self, text="Pressure", font="Helvetica 9 italic").grid(row=5, column=0, sticky='w')
        self.pLoad_ent = Entry(self, width=5, justify=CENTER)
        self.pLoad_ent.grid(row=5, column=1)
        self.pLoad_ent.insert(0, "1000")
        Label(self, text="kPa", font="Helvetica 9 italic").grid(row=5, column=2, sticky='w')

        # Number_of_steps
        Label(self, text="Steps", font="Helvetica 9 italic").grid(row=6, column=0, sticky='w')
        self.nst_ent = Entry(self, width=5, justify=CENTER)
        self.nst_ent.grid(row=6, column=1)
        self.nst_ent.insert(0, "10")

        tkinter.ttk.Separator(self, orient=HORIZONTAL).grid(column=0, row=7, columnspan=3, sticky='we')

        # Stamp width
        Label(self, text="Stamp width", font="Helvetica 9 italic").grid(row=8, column=0, sticky='w')
        self.b_ent = Entry(self, width=5, justify=CENTER)
        self.b_ent.grid(row=8, column=1)
        self.b_ent.insert(0, "2")
        Label(self, text="m", font="Helvetica 9 italic").grid(row=8, column=2, sticky='w')

        # Object width
        Label(self, text="Mesh width", font="Helvetica 9 italic").grid(row=9, column=0, sticky='w')
        self.ws_ent = Entry(self, width=5, justify=CENTER)
        self.ws_ent.grid(row=9, column=1)
        self.ws_ent.insert(0, "5")
        Label(self, text="m", font="Helvetica 9 italic").grid(row=9, column=2, sticky='w')

        # Object height
        Label(self, text="Mesh height", font="Helvetica 9 italic").grid(row=10, column=0, sticky='w')
        self.hs_ent = Entry(self, width=5, justify=CENTER)
        self.hs_ent.grid(row=10, column=1)
        self.hs_ent.insert(0, "4")
        Label(self, text="m", font="Helvetica 9 italic").grid(row=10, column=2, sticky='w')

        tkinter.ttk.Separator(self, orient=HORIZONTAL).grid(column=0, row=11, columnspan=3, sticky='we')

        # Width_number
        Label(self, text="Elements (Ox)", font="Helvetica 9 italic").grid(row=12, column=0, sticky='w')
        self.nw_ent = Entry(self, width=5, justify=CENTER)
        self.nw_ent.grid(row=12, column=1)
        self.nw_ent.insert(0, "20")

        # Height_number
        Label(self, text="Elements (Oz)", font="Helvetica 9 italic").grid(row=13, column=0, sticky='w')
        self.nh_ent = Entry(self, width=5, justify=CENTER)
        self.nh_ent.grid(row=13, column=1)
        self.nh_ent.insert(0, "16")

        # Stamp_number
        Label(self, text="Elements (stamp)", font="Helvetica 9 italic").grid(row=14, column=0, sticky='w')
        self.nb_ent = Entry(self, width=5, justify=CENTER)
        self.nb_ent.grid(row=14, column=1)
        self.nb_ent.insert(0, "12")

        tkinter.ttk.Separator(self, orient=HORIZONTAL).grid(column=0, row=15, columnspan=3, sticky='we')

        # Soil_model
        Label(self, text="Soil model:", font="Helvetica 9 italic").grid(row=16, column=0, sticky='w')
        # Plastic
        one = Radiobutton(self, text='Plastic', variable=self.radioValue, value=1, font="Helvetica 9 italic")
        one.grid(column=0, row=18, sticky="W")
        # Elastic
        two = Radiobutton(self, text='Elastic', variable=self.radioValue, value=0, font="Helvetica 9 italic")
        two.grid(column=0, row=17, sticky="W")

        # average
        aver = Checkbutton(self, text='Average', variable=self.var, font="Helvetica 9 italic")
        aver.grid(column=0, row=19, sticky="W")

        tkinter.ttk.Separator(self, orient=HORIZONTAL).grid(column=0, row=20, columnspan=3, sticky='we')

        # Btn_solve
        Button(self, text="Solve", command=self.solver, font="Helvetica 9 italic").grid(row=21, column=0,
                                                                                        columnspan=3, sticky='EW')

    def input(self):
        E = float(self.E_ent.get())
        nu = float(self.nu_ent.get())
        c = float(self.c_ent.get())
        fi = float(self.fi_ent.get())
        p_load = float(self.pLoad_ent.get())
        b = int(self.b_ent.get())
        ws = int(self.ws_ent.get())
        hs = int(self.hs_ent.get())
        nw = int(self.nw_ent.get())
        nh = int(self.nh_ent.get())
        nb = int(self.nb_ent.get())
        nst = int(self.nst_ent.get())
        plastic = self.radioValue.get()
        average = self.var.get()

        return E, nu, p_load, b, ws, hs, nw, nh, nb, c, fi, nst, plastic, average

    def solver(self):
        E, nu, p_load, b, ws, hs, nw, nh, nb, c, fi, nst, plastic, average = self.input()
        if plastic == 0:
            nst = 1
        # Mesh create
        b = b / 2
        # Nodes number
        n = (nw + 1) * (nh + 1)
        # Elements number
        nel = nw * nh * 2
        fi = fi * np.pi / 180
        niter = 20
        # Matrices create
        kx = np.zeros((nh + 1, nw + 1))
        kz = np.zeros((nh + 1, nw + 1))
        B = np.zeros((3, 6))
        D = np.zeros((3, 3))
        n_nmb = np.zeros((nh + 1, nw + 1))
        dex = np.zeros(nel)
        dez = np.zeros(nel)
        dexz = np.zeros(nel)
        ex = np.zeros((nst, nel))
        ez = np.zeros((nst, nel))
        exz = np.zeros((nst, nel))
        dsx = np.zeros(nel)
        dsz = np.zeros(nel)
        dtxz = np.zeros(nel)
        sx = np.zeros((nst, nel))
        sz = np.zeros((nst, nel))
        txz = np.zeros((nst, nel))
        s1 = np.zeros((nst, nel))
        s3 = np.zeros((nst, nel))
        pl = np.zeros(nel)
        Fp = np.zeros((nst, nel))
        FF = np.zeros((nst, nel, niter))
        npl = np.zeros((nst, niter))
        u = np.zeros((nst, n * 2))
        R = np.zeros((nst, n * 2))

        xk = 0
        zk = 0
        for i in range(nb):
            kx[0, i] = xk
            xk += b / nb
        for j in range(nb, nw + 1):
            kx[0, j] = xk
            xk += (ws - b) / (nw - nb)
        for k in range(1, nh + 1):
            kx[k] = kx[0]
        for i in range(nh + 1):
            kz[i, 0] = zk
            zk += hs / nh
        for i in range(1, nw + 1):
            for j in range(1, nh + 1):
                kz[j, i] = kz[j, i - 1]
        kxt = np.transpose(kx)
        kzt = np.transpose(kz)

        # Nodes numbers
        number = 1
        for i in range(nw + 1):
            for j in range(nh + 1):
                n_nmb[j, i] = number
                number += 1

        dp = p_load / nst
        for kst in range(0, nst + 1):
            if plastic == 0:
                kst = 1
            # Displacements_and_stresses_are_zero
            dR = np.zeros(n * 2)
            du = np.zeros(n * 2)

            for it in range(1, niter + 1):

                # Global_stiffness_matrix_is_zero
                kglob = np.zeros((n * 2, n * 2))

                # Stiffness matrix
                k = 0
                for i in range(1, nw + 1):
                    for j in range(1, nh + 1):
                        k = k + 1
                        l = (k + 1) // 2 + i - 1
                        ki = l
                        kj = l + 1
                        km = l + nh + 1
                        # Stiffness_matrix_element
                        xi = kx[j - 1, i - 1]
                        xj = kx[j, i - 1]
                        xm = kx[j - 1, i]
                        zi = kz[j - 1, i - 1]
                        zj = kz[j, i - 1]
                        zm = kz[j - 1, i]
                        BT = geometry_matrix_element(B, xi, xj, xm, zi, zj, zm)
                        D = elastic_matrix_element(D, E, nu)
                        D = plastic_matrix_element(k, it, sx, sz, txz, kst, nu, fi, D, pl, E)
                        kloc = stiffness_matrix_local(BT, D, B, xi, xj, xm, zi, zj, zm)
                        kglob = stiffness_matrix_total(kglob, ki, kj, kloc, km)

                        k = k + 1
                        m = (k + 2) // 2 + i - 1
                        ki = m
                        kj = m + nh
                        km = m + nh + 1
                        # Stiffness_matrix_element
                        xi = kx[j, i - 1]
                        xj = kx[j - 1, i]
                        xm = kx[j, i]
                        zi = kz[j, i - 1]
                        zj = kz[j - 1, i]
                        zm = kz[j, i]
                        BT = geometry_matrix_element(B, xi, xj, xm, zi, zj, zm)
                        D = elastic_matrix_element(D, E, nu)
                        D = plastic_matrix_element(k, it, sx, sz, txz, kst, nu, fi, D, pl, E)
                        kloc = stiffness_matrix_local(BT, D, B, xi, xj, xm, zi, zj, zm)
                        kglob = stiffness_matrix_total(kglob, ki, kj, kloc, km)

                # Boundary conditions
                for i in range(1, nb + 2):
                    j = ((i - 1) * nh + i) * 2
                    dR[j - 1] = dp * b / nb
                    if i == 1 or i == nb + 1:
                        dR[j - 1] = 0.5 * dp * b / nb

                for i in range(1, nh + 2):
                    j = i * 2
                    kglob[j - 2, j - 2] = 10 ** 16
                    j = (n - i + 1) * 2
                    kglob[j - 1, j - 1] = 10 ** 16
                    kglob[j - 2, j - 2] = 10 ** 16

                for i in range(nh + 1, n, nh + 1):
                    j = i * 2
                    kglob[j - 1, j - 1] = 10 ** 16

                # Solve SOLE
                du = np.dot(np.linalg.inv(kglob), dR)

                # Get Strains and deforms
                k = 0
                for i in range(1, nw + 1):
                    for j in range(1, nh + 1):
                        k = k + 1
                        l = (k + 1) // 2 + i - 1
                        ki = l
                        kj = l + 1
                        km = l + nh + 1
                        xi = kx[j - 1, i - 1]
                        xj = kx[j, i - 1]
                        xm = kx[j - 1, i]
                        zi = kz[j - 1, i - 1]
                        zj = kz[j, i - 1]
                        zm = kz[j - 1, i]
                        B = np.transpose(geometry_matrix_element(B, xi, xj, xm, zi, zj, zm))
                        D = elastic_matrix_element(D, E, nu)
                        D = plastic_matrix_element(k, it, sx, sz, txz, kst, nu, fi, D, pl, E)
                        dex, dez, dexz, ex, ez, exz = strains(du, B, ki, kj, km, k, dex, dez, dexz, kst, ex, ez, exz)
                        if average == 0:
                            dex, dez, dexz, sx, sz, txz, s1, s3 = stresses(D, dex, dez, dexz, k, dsx, dsz,
                                                                           dtxz, s1, s3, kst, sx, sz, txz)
                            sx, sz, txz, ex, ez, exz, dR, npl, Fp, FF, dex, dez, dexz = \
                                function_plastic(sx, sz, txz, kst, k, fi, c, it, Fp, FF, pl, npl,
                                                 nu, xi, xj, xm, zi, zj, zm, ki, kj, km, B, dR, dex,
                                                 dez, dexz, ex, ez, exz, E, plastic, average)

                        k = k + 1
                        m = (k + 2) // 2 + i - 1
                        ki = m
                        kj = m + nh
                        km = m + nh + 1
                        xi = kx[j, i - 1]
                        xj = kx[j - 1, i]
                        xm = kx[j, i]
                        zi = kz[j, i - 1]
                        zj = kz[j - 1, i]
                        zm = kz[j, i]
                        B = np.transpose(geometry_matrix_element(B, xi, xj, xm, zi, zj, zm))
                        D = elastic_matrix_element(D, E, nu)
                        D = plastic_matrix_element(k, it, sx, sz, txz, kst, nu, fi, D, pl, E)
                        dex, dez, dexz, ex, ez, exz = strains(du, B, ki, kj, km, k, dex, dez, dexz, kst, ex, ez, exz)
                        if average == 0:
                            dex, dez, dexz, sx, sz, txz, s1, s3 = stresses(D, dex, dez, dexz, k, dsx, dsz,
                                                                           dtxz, s1, s3, kst, sx, sz, txz)
                            sx, sz, txz, ex, ez, exz, dR, npl, Fp, FF, dex, dez, dexz = \
                                function_plastic(sx, sz, txz, kst, k, fi, c, it, Fp, FF, pl, npl,
                                                 nu, xi, xj, xm, zi, zj, zm, ki, kj, km, B, dR, dex,
                                                 dez, dexz, ex, ez, exz, E, plastic, average)
                        if average == 1:
                            dex, dez, dexz, sx, sz, txz, s1, s3 = stress_average(k, dex, dez, dexz, D, dsx,
                                                                                 dsz, dtxz, kst, sx, sz, txz, s1, s3)
                            sx, sz, txz, ex, ez, exz, dR, npl, Fp, FF, dex, dez, dexz = \
                                function_plastic(sx, sz, txz, kst, k, fi, c, it, Fp, FF, pl, npl,
                                                 nu, xi, xj, xm, zi, zj, zm, ki, kj, km, B, dR, dex,
                                                 dez, dexz, ex, ez, exz, E, plastic, average)

            for z in range(0, n * 2 + 1):
                if kst == 1:
                    u[kst - 1, z - 1] = du[z - 1]
                    R[kst - 1, z - 1] = dR[z - 1]
                else:
                    u[kst - 1, z - 1] = u[kst - 2, z - 1] + du[z - 1]
                    R[kst - 1, z - 1] = R[kst - 2, z - 1] + dR[z - 1]

        u_new = np.zeros(n)
        for i, j in zip(range(0, n * 2, 2), range(n + 1)):
            u_new[j] = u[kst - 1, i]
        u_new = np.transpose(np.reshape(u_new, (nw + 1, nh + 1)))
        kx_new = kx + u_new

        w_new = np.zeros(n)
        for i, j in zip(range(1, n * 2 + 1, 2), range(n + 1)):
            w_new[j] = u[kst - 1, i]
        w_new = np.transpose(np.reshape(w_new, (nw + 1, nh + 1)))
        kz_new = kz + w_new

        kxt_new = np.transpose(kx_new)
        kzt_new = np.transpose(kz_new)

        # Graphics
        ax.clear()
        ax.plot(kx, -kz, color='blue', linewidth=0.4)
        ax.plot(kxt, -kzt, color='blue', linewidth=0.4)
        for i in range(nh):
            for j in range(nw):
                ax.plot([kx[i + 1, j], kx[i, j + 1]], [-kz[i + 1, j], -kz[i, j + 1]], color='blue', linewidth=0.4)

        ax.plot(kx_new, -kz_new, color='red', linewidth=0.4)
        ax.plot(kxt_new, -kzt_new, color='red', linewidth=0.4)
        for i in range(nh):
            for j in range(nw):
                ax.plot([kx_new[i + 1, j], kx_new[i, j + 1]], [-kz_new[i + 1, j], -kz_new[i, j + 1]], color='red',
                        linewidth=0.4)

        for i in range(nh + 1):
            for j in range(nw + 1):
                ax.text(kx[i, j], -kz[i, j], n_nmb[i, j].astype(int), fontsize=5, fontstyle='italic')
        ax.xaxis.set_major_locator(ticker.MultipleLocator(0.5))
        ax.yaxis.set_major_locator(ticker.MultipleLocator(0.5))
        canvas.draw()
        canvas.get_tk_widget().place(x=175, y=0)

        # Table_creation
        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')
        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style='thin',
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style='thin',
                                     color='FF000000'),
                        vertical=Side(border_style='thin',
                                      color='FF000000'),
                        horizontal=Side(border_style='thin',
                                        color='FF000000')
                        )
        align_center = Alignment(horizontal='center',
                                 vertical='bottom',
                                 text_rotation=0,
                                 wrap_text=False,
                                 shrink_to_fit=False,
                                 indent=0)

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active
        ws.sheet_view.zoomScale = 85

        # название страницы
        ws.title = "Элементы"
        column = ['№ эл-та', 'sigma_x', 'sigma_z', 'tau_xz', 'sigma_1', 'sigma_3', 'eps_x', 'eps_z', 'eps_xz']
        for i, value in enumerate(column):
            ws.cell(row=1, column=i + 1).value = value
            ws.cell(row=1, column=i + 1).fill = fill
            ws.cell(row=1, column=i + 1).alignment = align_center
            ws.cell(row=1, column=i + 1).border = border
        for i in range(nel):
            ws.cell(row=i + 2, column=1).value = i + 1
            ws.cell(row=i + 2, column=1).alignment = align_center
            ws.cell(row=i + 2, column=1).border = border
            ws.cell(row=i + 2, column=2).value = sx[kst - 1, i]
            ws.cell(row=i + 2, column=2).alignment = align_center
            ws.cell(row=i + 2, column=2).border = border
            ws.cell(row=i + 2, column=2).number_format = '0.000'
            ws.cell(row=i + 2, column=3).value = sz[kst - 1, i]
            ws.cell(row=i + 2, column=3).alignment = align_center
            ws.cell(row=i + 2, column=3).border = border
            ws.cell(row=i + 2, column=3).number_format = '0.000'
            ws.cell(row=i + 2, column=4).value = txz[kst - 1, i]
            ws.cell(row=i + 2, column=4).alignment = align_center
            ws.cell(row=i + 2, column=4).border = border
            ws.cell(row=i + 2, column=4).number_format = '0.000'
            ws.cell(row=i + 2, column=5).value = s1[kst - 1, i]
            ws.cell(row=i + 2, column=5).alignment = align_center
            ws.cell(row=i + 2, column=5).border = border
            ws.cell(row=i + 2, column=5).number_format = '0.000'
            ws.cell(row=i + 2, column=6).value = s3[kst - 1, i]
            ws.cell(row=i + 2, column=6).alignment = align_center
            ws.cell(row=i + 2, column=6).border = border
            ws.cell(row=i + 2, column=6).number_format = '0.000'
            ws.cell(row=i + 2, column=7).value = ex[kst - 1, i]
            ws.cell(row=i + 2, column=7).alignment = align_center
            ws.cell(row=i + 2, column=7).border = border
            ws.cell(row=i + 2, column=7).number_format = '0.000'
            ws.cell(row=i + 2, column=8).value = ez[kst - 1, i]
            ws.cell(row=i + 2, column=8).alignment = align_center
            ws.cell(row=i + 2, column=8).border = border
            ws.cell(row=i + 2, column=8).number_format = '0.000'
            ws.cell(row=i + 2, column=9).value = exz[kst - 1, i]
            ws.cell(row=i + 2, column=9).alignment = align_center
            ws.cell(row=i + 2, column=9).border = border
            ws.cell(row=i + 2, column=9).number_format = '0.000'

        """# второй лист
        ws1 = wb.create_sheet("Матрица жесткости")
        ws1.sheet_view.zoomScale = 85
        for i in range(n * 2):
            for j in range(n * 2):
                ws1.cell(row=i + 1, column=j + 1).value = kglob[i, j]
                ws1.cell(row=i + 1, column=j + 1).alignment = align_center
                ws1.cell(row=i + 1, column=j + 1).border = border
                ws1.cell(row=i + 1, column=j + 1).number_format = '0.00'"""

        wb.save("Results.xlsx")


main_window = Tk()
main_window.title("FEM elastic plastic")
main_window.geometry("820x485")
fig = Figure()
ax = fig.add_subplot()
ax.xaxis.set_major_locator(ticker.MultipleLocator(0.5))
ax.yaxis.set_major_locator(ticker.MultipleLocator(0.5))
canvas = FigureCanvasTkAgg(fig, master=main_window)
canvas.draw()
toolbar = NavigationToolbar2Tk(canvas, main_window)
toolbar.update()
toolbar.place(x=175, y=0)
canvas.get_tk_widget().place(x=175, y=0)
app = Application(main_window)
main_window.mainloop()
